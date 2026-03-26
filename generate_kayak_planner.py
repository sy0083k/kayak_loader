"""
카약 캠핑 적재 계획 수립 도우미 — Excel 파일 생성기
생성: kayak_packing_planner.xlsx  +  kayak_vba.bas
2대 카약 지원 + ActiveX 체크박스 LinkedCell VBA 연동
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ── 공통 색상 ──────────────────────────────────────────────────────────────
CLR = {
    "yellow_input":  "FFFFE0",
    "gray_label":    "D9D9D9",
    "blue_header":   "1565C0",
    "white":         "FFFFFF",
    "light_blue":    "DDEEFF",
    "bow_bg":        "E0F7FA",   # 선수 (카약1)
    "stern_bg":      "FFF3E0",   # 선미 (카약1)
    "k2_bow":        "E8F5E9",   # 카약2-선수 연두
    "k2_stern":      "FFF9C4",   # 카약2-선미 연노랑
    "unassigned":    "F5F5F5",   # 미배정 회색
    "red_warn":      "FFE0E0",
    "green_ok":      "C8E6C9",
    "orange_caution":"FFE0B2",
    "red_bad":       "FFCDD2",
    "dark_bg":       "1A237E",
    "title_blue":    "0D47A1",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border():
    s = Side(style="medium")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_border(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1: 카약_설정  (카약 1 / 카약 2 나란히)
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet_settings(wb):
    ws = wb.create_sheet("카약_설정")
    # 열 너비: A(4) B(26) C(18) D(3 구분) E(26) F(18) G(24 힌트)
    col_widths = [4, 26, 18, 3, 26, 18, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 제목 행
    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value = "🛶 카약 설정 (카약 1 / 카약 2)"
    c.font = Font(bold=True, size=16, color=CLR["white"])
    c.fill = fill(CLR["title_blue"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 카약 1 / 카약 2 구분 헤더
    ws.row_dimensions[2].height = 22
    ws.merge_cells("B2:C2")
    k1h = ws["B2"]
    k1h.value = "── 카약 1 ──"
    k1h.font = Font(bold=True, color=CLR["white"], size=11)
    k1h.fill = fill("1565C0")
    k1h.alignment = Alignment(horizontal="center", vertical="center")
    k1h.border = thin_border()

    ws.merge_cells("E2:F2")
    k2h = ws["E2"]
    k2h.value = "── 카약 2 ──"
    k2h.font = Font(bold=True, color=CLR["white"], size=11)
    k2h.fill = fill("E65100")
    k2h.alignment = Alignment(horizontal="center", vertical="center")
    k2h.border = thin_border()

    def label_cell(row, col, text, bg=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(bold=True, size=11)
        c.fill = fill(bg or CLR["gray_label"])
        c.alignment = Alignment(vertical="center")
        c.border = thin_border()
        return c

    def input_cell(row, col, hint=""):
        c = ws.cell(row=row, column=col)
        c.fill = fill(CLR["yellow_input"])
        c.border = thin_border()
        c.alignment = Alignment(horizontal="left", vertical="center")
        if hint:
            d = ws.cell(row=row, column=7, value=hint)
            d.font = Font(italic=True, color="888888", size=10)
            d.alignment = Alignment(vertical="center")
        return c

    for r in range(3, 12):
        ws.row_dimensions[r].height = 22

    # 기종명 (row 3)
    label_cell(3, 2, "카약 기종명")
    input_cell(3, 3, "예: Valley Etain 17.7")
    label_cell(3, 5, "카약 기종명")
    input_cell(3, 6)

    # 선수 해치 구분 헤더 (row 4)
    label_cell(4, 2, "── 선수(Bow) 해치 ──", "B3E5FC")
    label_cell(4, 5, "── 선수(Bow) 해치 ──", "B3E5FC")

    # 선수 최대 중량 (row 5)
    label_cell(5, 2, "  최대 적재 중량 (kg)")
    input_cell(5, 3, "예: 25")
    label_cell(5, 5, "  최대 적재 중량 (kg)")
    input_cell(5, 6)

    # 선수 최대 부피 (row 6)
    label_cell(6, 2, "  최대 적재 용적 (L)")
    input_cell(6, 3, "예: 60")
    label_cell(6, 5, "  최대 적재 용적 (L)")
    input_cell(6, 6)

    # 선미 해치 구분 헤더 (row 7)
    label_cell(7, 2, "── 선미(Stern) 해치 ──", "FFE0B2")
    label_cell(7, 5, "── 선미(Stern) 해치 ──", "FFE0B2")

    # 선미 최대 중량 (row 8)
    label_cell(8, 2, "  최대 적재 중량 (kg)")
    input_cell(8, 3)
    label_cell(8, 5, "  최대 적재 중량 (kg)")
    input_cell(8, 6)

    # 선미 최대 부피 (row 9)
    label_cell(9, 2, "  최대 적재 용적 (L)")
    input_cell(9, 3)
    label_cell(9, 5, "  최대 적재 용적 (L)")
    input_cell(9, 6)

    # 패들러 정보 헤더 (row 10)
    label_cell(10, 2, "── 패들러 정보 ──", "F3E5F5")
    label_cell(10, 5, "── 패들러 정보 ──", "F3E5F5")

    # 패들러 체중 (row 11)
    label_cell(11, 2, "  패들러 체중 (kg)")
    input_cell(11, 3, "균형 계산 참고용")
    label_cell(11, 5, "  패들러 체중 (kg)")
    input_cell(11, 6)

    # 기본값
    ws["C3"] = "Valley Etain 17.7"
    ws["C5"] = 25
    ws["C6"] = 60
    ws["C8"] = 30
    ws["C9"] = 80
    ws["C11"] = 75

    ws["F3"] = "Necky Chatham 17"
    ws["F5"] = 25
    ws["F6"] = 60
    ws["F8"] = 30
    ws["F9"] = 80
    ws["F11"] = 70


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2: 장비_DB  (H/I/J/K = ActiveX 체크박스 LinkedCell)
# ══════════════════════════════════════════════════════════════════════════════
EQUIPMENT_DB = [
    # (장비명, 카테고리, 무게g, 부피L, 필수여부, 비고)
    ("경량 텐트",        "취침", 1800, 4.5, "필수", "1~2인용 자립형"),
    ("침낭 (3계절)",     "취침", 900,  3.0, "필수", "방수 압축 팩 필수"),
    ("슬리핑 패드",      "취침", 450,  2.5, "필수", "접이식 폼 또는 에어"),
    ("알코올 버너",      "식사", 150,  0.3, "필수", "MSR Pocket Rocket 류"),
    ("가스 캐니스터",    "식사", 400,  0.5, "필수", "230g 캔 기준"),
    ("코펠 세트",        "식사", 350,  0.8, "필수", "티타늄 1~2인용"),
    ("식기 세트",        "식사", 120,  0.3, "선택", "스푼·포크·젓가락"),
    ("비상 식량 1일분",  "식사", 600,  1.2, "필수", "에너지바·건조식품"),
    ("물 (2L)",          "식사", 2000, 2.0, "필수", "정수 필터 포함 가능"),
    ("방수 재킷",        "의류", 500,  1.5, "필수", "고어텍스 또는 동급"),
    ("여벌 패들링복",    "의류", 300,  0.8, "선택", "드라이 또는 웻수트"),
    ("모자 (챙 넓은)",   "의류", 100,  0.4, "필수", "자외선 차단용"),
    ("샌들·슬리퍼",     "의류", 300,  0.6, "선택", "상륙 시 착용"),
    ("구명조끼 (PFD)",   "안전", 1200, 3.0, "필수", "항시 착용 필수"),
    ("구급 키트",        "안전", 250,  0.5, "필수", "방수 케이스 포장"),
    ("호루라기",         "안전", 20,   0.0, "필수", "PFD에 부착"),
    ("방수 라이트",      "안전", 150,  0.2, "필수", "야간 야영·신호용"),
    ("스프레이 스커트",  "패들링", 500, 1.0, "필수", "카약 콕핏 크기 확인"),
    ("빌지 펌프",        "패들링", 300, 0.8, "필수", "핸드 펌프 타입"),
    ("패들 플로트",      "패들링", 350, 2.5, "필수", "재진입 구조용"),
    ("드라이백 20L",     "기타",  200,  0.3, "필수", "방수 롤탑"),
    ("드라이백 5L",      "기타",  100,  0.1, "선택", "소형 장비 분류용"),
    ("액션 카메라",      "기타",  120,  0.2, "선택", "방수 케이스 포함"),
    ("선크림 SPF50+",    "기타",   80,  0.1, "필수", "물에 지워지지 않는"),
    ("내비게이션 지도",  "기타",   50,  0.1, "선택", "방수 케이스에 보관"),

    # ── 취침 추가 (3)
    ("텐트 그라운드시트", "취침",  300,  0.5, "선택", "텐트 바닥 보호·방수"),
    ("침낭 라이너",       "취침",  200,  0.4, "선택", "보온 강화·위생 관리"),
    ("에어 베개",         "취침",   80,  0.2, "선택", "팽창식 초경량"),

    # ── 식사 추가 (4)
    ("정수 필터",         "식사",  100,  0.2, "선택", "소이어 스퀴즈 류"),
    ("조미료 세트",       "식사",  150,  0.3, "선택", "소금·후추·참기름 소분"),
    ("식품 보관백",       "식사",   50,  0.1, "선택", "냄새 차단 지퍼백"),
    ("커피·차 용품",      "식사",  120,  0.3, "선택", "드립백·티백 세트"),

    # ── 의류 추가 (4)
    ("방수 장갑",         "의류",  180,  0.3, "선택", "네오프렌 또는 합성 소재"),
    ("래쉬가드",          "의류",  200,  0.5, "선택", "자외선 차단 장·반팔"),
    ("방한 플리스",       "의류",  400,  1.0, "선택", "상륙 후 체온 유지"),
    ("방수 바지",         "의류",  350,  0.7, "선택", "상·하반신 레이어링"),

    # ── 안전 추가 (3)
    ("토우라인",          "안전",  400,  0.8, "필수", "15m 이상 부유 구조로프"),
    ("신호 거울",         "안전",   50,  0.0, "선택", "조난 신호용"),
    ("케미라이트",        "안전",   30,  0.0, "선택", "야간 시인성·비상신호"),

    # ── 패들링 추가 (4)
    ("나침반",            "패들링",  80,  0.1, "필수", "방수 독립형 나침반"),
    ("패들 리쉬",         "패들링",  60,  0.1, "선택", "패들 분실 방지 줄"),
    ("씨앵커",            "패들링", 250,  0.5, "선택", "드리프트 제어 파라슈트"),
    ("카약 카트",         "패들링",1500,  3.0, "선택", "해변·선착장 이동 트롤리"),

    # ── 기타 추가 (7)
    ("헤드램프",          "기타",  100,  0.2, "필수", "방수 IPX4 이상"),
    ("보조 배터리",       "기타",  350,  0.5, "선택", "방수 케이스·10000mAh"),
    ("방수 폰케이스",     "기타",   50,  0.1, "필수", "터치 조작 가능 타입"),
    ("접이식 의자",       "기타",  800,  2.0, "선택", "경량 캠핑 체어"),
    ("쓰레기 봉투",       "기타",   50,  0.1, "필수", "LNT — 흔적 남기지 않기"),
    ("모기 기피제",       "기타",  100,  0.1, "선택", "펌프 또는 롤온 타입"),
    ("수리 키트",         "기타",  150,  0.2, "선택", "텐트·건조수트 응급 수리"),
]

DB_ROW_COUNT = len(EQUIPMENT_DB)   # 50

def build_sheet_db(wb):
    ws = wb.create_sheet("장비_DB")

    headers = ["번호", "장비명", "카테고리", "무게(g)", "부피(L)", "필수여부", "비고",
               "카약1_선수✓", "카약1_선미✓", "카약2_선수✓", "카약2_선미✓"]
    col_widths = [6, 22, 12, 12, 10, 10, 20, 12, 12, 12, 12]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 헤더 색상 설정
    header_colors = {
        1: CLR["blue_header"], 2: CLR["blue_header"], 3: CLR["blue_header"],
        4: CLR["blue_header"], 5: CLR["blue_header"], 6: CLR["blue_header"],
        7: CLR["blue_header"],
        8:  "00838F",   # 카약1-선수: 청록
        9:  "4DB6AC",   # 카약1-선미: 연청록
        10: "E65100",   # 카약2-선수: 주황
        11: "FF8A65",   # 카약2-선미: 연주황
    }

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=CLR["white"], size=11)
        c.fill = fill(header_colors[col])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[1].height = 24

    alt_fill = fill("DDEEFF")
    white_fill = fill(CLR["white"])

    for i, (name, cat, weight, vol, req, note) in enumerate(EQUIPMENT_DB, 1):
        row = i + 1
        row_fill = white_fill if i % 2 == 0 else alt_fill

        data = [i, name, cat, weight, vol, req, note]
        for col, val in enumerate(data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border()
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (1, 3, 6) else "left")
            c.fill = row_fill
            ws.row_dimensions[row].height = 24

        # H~K: ActiveX 체크박스 LinkedCell 대상 (초기값 FALSE)
        for col in range(8, 12):
            c = ws.cell(row=row, column=col, value=False)
            c.fill = fill(CLR["yellow_input"])
            c.border = thin_border()
            c.alignment = Alignment(horizontal="center", vertical="center")

    # 숫자 서식
    for row in range(2, DB_ROW_COUNT + 2):
        ws.cell(row=row, column=4).number_format = '#,##0 "g"'
        ws.cell(row=row, column=5).number_format = '0.0 "L"'


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3: 적재_계획
# ══════════════════════════════════════════════════════════════════════════════
PLAN_DATA_START = 4
PLAN_DATA_END   = PLAN_DATA_START + DB_ROW_COUNT - 1   # = 28

def build_sheet_plan(wb):
    ws = wb.create_sheet("적재_계획")

    # 열 너비: A B C D E F G H I J
    col_widths = [4, 26, 14, 11, 9, 12, 12, 14, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 섹션 제목
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "🗂 적재 계획 (카약 1 / 카약 2)"
    c.font = Font(bold=True, size=14, color=CLR["white"])
    c.fill = fill(CLR["title_blue"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # ── 범례
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = (
        "  카약1: 선수=청록(E0F7FA) / 선미=주황(FFF3E0)   "
        "카약2: 선수=연두(E8F5E9) / 선미=연노랑(FFF9C4)   "
        "미배정=회색(F5F5F5)"
    )
    c.font = Font(italic=True, size=9, color="444444")
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 16

    # ── 헤더 행
    headers = ["#", "장비명", "카테고리", "무게(g)", "부피(L)",
               "카약1배치", "카약2배치", "포장방법", "우선순위", "비고"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color=CLR["white"], size=10)
        c.fill = fill(CLR["blue_header"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[3].height = 22

    # 데이터 유효성: 포장방법 / 우선순위 (F/G는 VBA가 쓰므로 DV 없음)
    dv_packing = DataValidation(
        type="list",
        formula1='"드라이백,직접적재,메쉬백"',
        allow_blank=True,
        showDropDown=False
    )
    dv_priority = DataValidation(
        type="list",
        formula1='"상,중,하"',
        allow_blank=True,
        showDropDown=False
    )
    ws.add_data_validation(dv_packing)
    ws.add_data_validation(dv_priority)

    for i in range(DB_ROW_COUNT):
        row    = PLAN_DATA_START + i   # 4~28
        db_row = i + 2                 # DB 행: 2~26

        ws.row_dimensions[row].height = 20

        # A: 번호 (정적)
        a = ws.cell(row=row, column=1, value=i + 1)
        a.alignment = Alignment(horizontal="center", vertical="center")
        a.border = thin_border()
        a.fill = fill("F5F5F5")

        # B: 장비명 — 장비_DB 직접 참조
        b = ws.cell(row=row, column=2, value=f"=장비_DB!B{db_row}")
        b.border = thin_border()
        b.alignment = Alignment(vertical="center")

        # C: 카테고리
        cc = ws.cell(row=row, column=3, value=f"=장비_DB!C{db_row}")
        cc.border = thin_border()
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.fill = fill("F0F4FF")

        # D: 무게(g)
        d = ws.cell(row=row, column=4, value=f"=장비_DB!D{db_row}")
        d.border = thin_border()
        d.number_format = '#,##0'
        d.alignment = Alignment(horizontal="right", vertical="center")
        d.fill = fill("F0F4FF")

        # E: 부피(L)
        e = ws.cell(row=row, column=5, value=f"=장비_DB!E{db_row}")
        e.border = thin_border()
        e.number_format = '0.0'
        e.alignment = Alignment(horizontal="right", vertical="center")
        e.fill = fill("F0F4FF")

        # F: 카약1 배치 — 장비_DB H/I 체크박스 LinkedCell 직접 참조
        # Form Control은 Worksheet_Change를 트리거하지 않으므로 수식으로 자동반영
        f = ws.cell(row=row, column=6,
                    value=f'=IF(장비_DB!H{db_row},"선수",IF(장비_DB!I{db_row},"선미",""))')
        f.border = thin_border()
        f.alignment = Alignment(horizontal="center", vertical="center")

        # G: 카약2 배치 — 장비_DB J/K 체크박스 LinkedCell 직접 참조
        g = ws.cell(row=row, column=7,
                    value=f'=IF(장비_DB!J{db_row},"선수",IF(장비_DB!K{db_row},"선미",""))')
        g.border = thin_border()
        g.alignment = Alignment(horizontal="center", vertical="center")

        # H: 포장방법 드롭다운
        h = ws.cell(row=row, column=8)
        h.border = thin_border()
        h.alignment = Alignment(horizontal="center", vertical="center")
        dv_packing.add(h)

        # I: 우선순위 드롭다운
        ii = ws.cell(row=row, column=9)
        ii.border = thin_border()
        ii.alignment = Alignment(horizontal="center", vertical="center")
        dv_priority.add(ii)

        # J: 비고
        j = ws.cell(row=row, column=10)
        j.border = thin_border()

    # ── 조건부 서식 (우선순위: 나중 추가 = 높은 순위)
    cf_range = f"A{PLAN_DATA_START}:J{PLAN_DATA_END}"
    S = PLAN_DATA_START

    # 1. 미배정 (낮은 순위, 먼저 추가)
    ws.conditional_formatting.add(cf_range,
        FormulaRule(formula=[f'AND($F{S}="",$G{S}="")'],
                    fill=fill(CLR["unassigned"])))
    # 2. 카약2 선미
    ws.conditional_formatting.add(cf_range,
        FormulaRule(formula=[f'$G{S}="선미"'], fill=fill(CLR["k2_stern"])))
    # 3. 카약2 선수
    ws.conditional_formatting.add(cf_range,
        FormulaRule(formula=[f'$G{S}="선수"'], fill=fill(CLR["k2_bow"])))
    # 4. 카약1 선미
    ws.conditional_formatting.add(cf_range,
        FormulaRule(formula=[f'$F{S}="선미"'], fill=fill(CLR["stern_bg"])))
    # 5. 카약1 선수 (높은 순위, 마지막 추가)
    ws.conditional_formatting.add(cf_range,
        FormulaRule(formula=[f'$F{S}="선수"'], fill=fill(CLR["bow_bg"])))

    # ── 요약 섹션
    # 섹션 제목
    title_row = PLAN_DATA_END + 1   # 29
    ws.merge_cells(f"A{title_row}:J{title_row}")
    t = ws.cell(row=title_row, column=1, value="⚖ 무게 & 부피 균형 요약")
    t.font = Font(bold=True, size=13, color=CLR["white"])
    t.fill = fill("37474F")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[title_row].height = 26

    summary_start = PLAN_DATA_END + 2   # 30

    dr_d = f"D{PLAN_DATA_START}:D{PLAN_DATA_END}"
    dr_e = f"E{PLAN_DATA_START}:E{PLAN_DATA_END}"
    dr_f = f"F{PLAN_DATA_START}:F{PLAN_DATA_END}"
    dr_g = f"G{PLAN_DATA_START}:G{PLAN_DATA_END}"

    # 요약 컬럼 헤더 (A=라벨, B=카약1선수, C=카약1선미, D=카약2선수, E=카약2선미)
    hdr_row = summary_start - 1   # 29 — but we used that for the title
    # Insert column headers as part of title row or skip — keep title at 29
    # Column headers will be implicit through row labels

    def sum_label(row, text, bg="ECEFF1"):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=10)
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 20

    def sum_val(row, col, formula, num_fmt=None, bg=None):
        c = ws.cell(row=row, column=col, value=formula)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="right", vertical="center")
        if num_fmt:
            c.number_format = num_fmt
        if bg:
            c.fill = fill(bg)
        return c

    # 소제목 행 (B/C/D/E 그룹 헤더)
    sub_hdr_row = summary_start - 1   # 29 is already used by section title
    # We'll put group headers at summary_start, shift data by 1
    # Actually let's put group label row at summary_start, data from summary_start+1
    grp_hdr = summary_start   # 30
    for col, txt, bg in [
        (2, "카약1 선수", "E0F7FA"), (3, "카약1 선미", "FFF3E0"),
        (4, "카약2 선수", "E8F5E9"), (5, "카약2 선미", "FFF9C4"),
    ]:
        c = ws.cell(row=grp_hdr, column=col, value=txt)
        c.font = Font(bold=True, size=9, color="333333")
        c.fill = fill(bg)
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[grp_hdr].height = 18

    r = summary_start + 1   # 31  ← weight_row

    # 적재 중량 합계 (kg)
    sum_label(r, "적재 중량 합계 (kg)")
    sum_val(r, 2, f'=SUMIF({dr_f},"선수",{dr_d})/1000', '0.00 "kg"')
    sum_val(r, 3, f'=SUMIF({dr_f},"선미",{dr_d})/1000', '0.00 "kg"')
    sum_val(r, 4, f'=SUMIF({dr_g},"선수",{dr_d})/1000', '0.00 "kg"')
    sum_val(r, 5, f'=SUMIF({dr_g},"선미",{dr_d})/1000', '0.00 "kg"')
    weight_row = r

    r += 1   # 32
    # 최대 허용 중량
    sum_label(r, "최대 허용 중량 (kg)", "E3F2FD")
    sum_val(r, 2, "=카약_설정!C5", '0.0 "kg"')
    sum_val(r, 3, "=카약_설정!C8", '0.0 "kg"')
    sum_val(r, 4, "=카약_설정!F5", '0.0 "kg"')
    sum_val(r, 5, "=카약_설정!F8", '0.0 "kg"')

    r += 1   # 33
    # 중량 사용률
    sum_label(r, "중량 사용률 (%)")
    for col, max_ref in [(2, "카약_설정!C5"), (3, "카약_설정!C8"),
                         (4, "카약_설정!F5"), (5, "카약_설정!F8")]:
        wc = get_column_letter(col)
        formula = f"=IF({max_ref}>0,{wc}{weight_row}/{max_ref},\"\")"
        cv = sum_val(r, col, formula, '0.0%')
        cell_ref = f"{wc}{r}"
        ws.conditional_formatting.add(cell_ref,
            CellIsRule(operator="greaterThan", formula=["1"], fill=fill(CLR["red_warn"])))
    weight_pct_row = r

    r += 1   # 34
    sum_label(r, "─── 부피 ───", "E8F5E9")

    r += 1   # 35
    # 부피 합계
    sum_label(r, "적재 부피 합계 (L)")
    sum_val(r, 2, f'=SUMIF({dr_f},"선수",{dr_e})', '0.0 "L"')
    sum_val(r, 3, f'=SUMIF({dr_f},"선미",{dr_e})', '0.0 "L"')
    sum_val(r, 4, f'=SUMIF({dr_g},"선수",{dr_e})', '0.0 "L"')
    sum_val(r, 5, f'=SUMIF({dr_g},"선미",{dr_e})', '0.0 "L"')
    vol_row = r

    r += 1   # 36
    # 최대 허용 부피
    sum_label(r, "최대 허용 부피 (L)", "E3F2FD")
    sum_val(r, 2, "=카약_설정!C6", '0.0 "L"')
    sum_val(r, 3, "=카약_설정!C9", '0.0 "L"')
    sum_val(r, 4, "=카약_설정!F6", '0.0 "L"')
    sum_val(r, 5, "=카약_설정!F9", '0.0 "L"')

    r += 1   # 37
    # 부피 사용률
    sum_label(r, "부피 사용률 (%)")
    for col, max_ref in [(2, "카약_설정!C6"), (3, "카약_설정!C9"),
                         (4, "카약_설정!F6"), (5, "카약_설정!F9")]:
        wc = get_column_letter(col)
        formula = f"=IF({max_ref}>0,{wc}{vol_row}/{max_ref},\"\")"
        sum_val(r, col, formula, '0.0%')
        ws.conditional_formatting.add(f"{wc}{r}",
            CellIsRule(operator="greaterThan", formula=["1"], fill=fill(CLR["red_warn"])))
    vol_pct_row = r

    r += 2   # 39 — skip one row

    # 카약1 균형 판정 (B:C merge)
    k1_balance_row = r
    ws.row_dimensions[r].height = 26
    lc1 = ws.cell(row=r, column=1, value="카약1 균형")
    lc1.font = Font(bold=True, color=CLR["white"])
    lc1.fill = fill("37474F")
    lc1.border = thin_border()
    lc1.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"B{r}:C{r}")
    bc1 = ws.cell(row=r, column=2)
    bw1 = f"B{weight_row}"; sw1 = f"C{weight_row}"
    bc1.value = (
        f'=IF(OR({bw1}+{sw1}=0,{bw1}+{sw1}=""),"(데이터 없음)",'
        f'IF(AND({bw1}/({bw1}+{sw1})>=0.4,{bw1}/({bw1}+{sw1})<=0.6),"✅ 양호",'
        f'IF(AND({bw1}/({bw1}+{sw1})>=0.3,{bw1}/({bw1}+{sw1})<=0.7),"⚠️ 주의","❌ 불균형")))'
    )
    bc1.font = Font(bold=True, size=13)
    bc1.alignment = Alignment(horizontal="center", vertical="center")
    bc1.border = thin_border()

    for ref in [f"B{r}:C{r}"]:
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$B{r}="✅ 양호"'], fill=fill(CLR["green_ok"])))
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$B{r}="⚠️ 주의"'], fill=fill(CLR["orange_caution"])))
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$B{r}="❌ 불균형"'], fill=fill(CLR["red_bad"])))

    r += 1   # 40

    # 카약2 균형 판정 (D:E merge)
    k2_balance_row = r
    ws.row_dimensions[r].height = 26
    lc2 = ws.cell(row=r, column=1, value="카약2 균형")
    lc2.font = Font(bold=True, color=CLR["white"])
    lc2.fill = fill("37474F")
    lc2.border = thin_border()
    lc2.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"D{r}:E{r}")
    bc2 = ws.cell(row=r, column=4)
    bw2 = f"D{weight_row}"; sw2 = f"E{weight_row}"
    bc2.value = (
        f'=IF(OR({bw2}+{sw2}=0,{bw2}+{sw2}=""),"(데이터 없음)",'
        f'IF(AND({bw2}/({bw2}+{sw2})>=0.4,{bw2}/({bw2}+{sw2})<=0.6),"✅ 양호",'
        f'IF(AND({bw2}/({bw2}+{sw2})>=0.3,{bw2}/({bw2}+{sw2})<=0.7),"⚠️ 주의","❌ 불균형")))'
    )
    bc2.font = Font(bold=True, size=13)
    bc2.alignment = Alignment(horizontal="center", vertical="center")
    bc2.border = thin_border()

    for ref in [f"D{r}:E{r}"]:
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$D{r}="✅ 양호"'], fill=fill(CLR["green_ok"])))
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$D{r}="⚠️ 주의"'], fill=fill(CLR["orange_caution"])))
        ws.conditional_formatting.add(ref,
            FormulaRule(formula=[f'$D{r}="❌ 불균형"'], fill=fill(CLR["red_bad"])))

    r += 1   # 41 — 카약1 선수:선미 비율
    ws.row_dimensions[r].height = 20
    lratio1 = ws.cell(row=r, column=1, value="선수:선미 비율 (카약1)")
    lratio1.font = Font(bold=True)
    lratio1.fill = fill("ECEFF1")
    lratio1.border = thin_border()
    ws.merge_cells(f"B{r}:C{r}")
    rv1 = ws.cell(row=r, column=2)
    rv1.value = (
        f'=IF(B{weight_row}+C{weight_row}=0,"",'
        f'TEXT(B{weight_row}/(B{weight_row}+C{weight_row}),"0%")&" : "'
        f'&TEXT(C{weight_row}/(B{weight_row}+C{weight_row}),"0%"))'
    )
    rv1.alignment = Alignment(horizontal="center", vertical="center")
    rv1.border = thin_border()

    r += 1   # 42 — 카약2 선수:선미 비율
    ws.row_dimensions[r].height = 20
    lratio2 = ws.cell(row=r, column=1, value="선수:선미 비율 (카약2)")
    lratio2.font = Font(bold=True)
    lratio2.fill = fill("ECEFF1")
    lratio2.border = thin_border()
    ws.merge_cells(f"D{r}:E{r}")
    rv2 = ws.cell(row=r, column=4)
    rv2.value = (
        f'=IF(D{weight_row}+E{weight_row}=0,"",'
        f'TEXT(D{weight_row}/(D{weight_row}+E{weight_row}),"0%")&" : "'
        f'&TEXT(E{weight_row}/(D{weight_row}+E{weight_row}),"0%"))'
    )
    rv2.alignment = Alignment(horizontal="center", vertical="center")
    rv2.border = thin_border()

    return {
        "weight_row":     weight_row,
        "weight_pct_row": weight_pct_row,
        "vol_row":        vol_row,
        "vol_pct_row":    vol_pct_row,
        "k1_balance_row": k1_balance_row,
        "k2_balance_row": k2_balance_row,
        "data_start":     PLAN_DATA_START,
        "data_end":       PLAN_DATA_END,
    }


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 4: 대시보드  (2대 카약 × 4그룹)
# ══════════════════════════════════════════════════════════════════════════════
def build_sheet_dashboard(wb, plan_refs):
    ws = wb.create_sheet("대시보드")

    # 13열 × 13pt
    for i in range(1, 14):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.column_dimensions["A"].width = 5

    # ── 1. 메인 타이틀 (B1:M2)
    ws.merge_cells("B1:M2")
    t = ws["B1"]
    t.value = "🛶  카약 적재 계획 대시보드 (2대)"
    t.font = Font(bold=True, size=20, color=CLR["white"])
    t.fill = fill(CLR["dark_bg"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 28

    # ── 2. 카약1 / 카약2 정보 박스
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22

    # 카약1 박스 (B~G = cols 2~7), 카약2 박스 (H~M = cols 8~13)
    # 각 박스: 3쌍 (각 2열) = 기종명, 선수최대, 선미최대
    kayak_info = [
        # (col_start, hdr_color, kayak_label, name_ref, bow_w_ref, bow_v_ref, stern_w_ref, stern_v_ref)
        (2, "1565C0", "카약 1",
         "=카약_설정!C3", '=카약_설정!C5&" kg"', '=카약_설정!C6&" L"',
         '=카약_설정!C8&" kg"', '=카약_설정!C9&" L"'),
        (8, "E65100", "카약 2",
         "=카약_설정!F3", '=카약_설정!F5&" kg"', '=카약_설정!F6&" L"',
         '=카약_설정!F8&" kg"', '=카약_설정!F9&" L"'),
    ]

    for col_start, hdr_color, kayak_label, name_ref, bw_ref, bv_ref, sw_ref, sv_ref in kayak_info:
        # Kayak label spanning header
        ws.merge_cells(start_row=3, end_row=3,
                       start_column=col_start, end_column=col_start + 5)
        hdr = ws.cell(row=3, column=col_start, value=kayak_label)
        hdr.font = Font(bold=True, color=CLR["white"], size=12)
        hdr.fill = fill(hdr_color)
        hdr.alignment = Alignment(horizontal="center", vertical="center")
        hdr.border = thin_border()
        ws.row_dimensions[3].height = 22

        # 선수/선미 최대값: 두 참조를 하나의 유효한 Excel 연결 수식으로 합성
        # bw_ref 예: '=카약_설정!C5&" kg"'  → [1:] 로 '=' 제거 후 연결
        bow_combined   = f'={bw_ref[1:]}&" / "&{bv_ref[1:]}'
        stern_combined = f'={sw_ref[1:]}&" / "&{sv_ref[1:]}'
        short_pairs = [
            ("기종명",              name_ref),
            ("선수 최대(중량/부피)", bow_combined),
            ("선미 최대(중량/부피)", stern_combined),
        ]

        for idx, (lbl, formula) in enumerate(short_pairs):
            col = col_start + idx * 2
            ws.merge_cells(start_row=4, end_row=4,
                           start_column=col, end_column=col + 1)
            lc = ws.cell(row=4, column=col, value=lbl)
            lc.font = Font(bold=True, color=CLR["white"], size=9)
            lc.fill = fill(hdr_color)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border = thin_border()

            ws.merge_cells(start_row=5, end_row=5,
                           start_column=col, end_column=col + 1)
            vc = ws.cell(row=5, column=col, value=formula)
            vc.font = Font(bold=True, size=10)
            vc.fill = fill("ECEFF1")
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = thin_border()

    # ── 3. 중량 & 부피 현황 섹션
    ws.row_dimensions[7].height = 24
    ws.merge_cells("B7:M7")
    s2 = ws["B7"]
    s2.value = "중량 & 부피 현황"
    s2.font = Font(bold=True, color=CLR["white"], size=12)
    s2.fill = fill("455A64")
    s2.alignment = Alignment(horizontal="center", vertical="center")

    wr  = plan_refs["weight_row"]
    wpr = plan_refs["weight_pct_row"]
    vr  = plan_refs["vol_row"]
    vpr = plan_refs["vol_pct_row"]

    # 4그룹: K1선수(B), K1선미(C), K2선수(D), K2선미(E) in 적재_계획
    groups = [
        ("카약1 선수", f"=적재_계획!B{wr}", f"=적재_계획!B{wpr}", "E0F7FA"),
        ("카약1 선미", f"=적재_계획!C{wr}", f"=적재_계획!C{wpr}", "FFF3E0"),
        ("카약2 선수", f"=적재_계획!D{wr}", f"=적재_계획!D{wpr}", "E8F5E9"),
        ("카약2 선미", f"=적재_계획!E{wr}", f"=적재_계획!E{wpr}", "FFF9C4"),
    ]

    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 26
    ws.row_dimensions[10].height = 20
    ws.row_dimensions[11].height = 26

    for idx, (lbl, w_formula, pct_formula, bg) in enumerate(groups):
        col = 2 + idx * 3  # B=2, E=5, H=8, K=11

        # 중량 수치 라벨 (row 8)
        ws.merge_cells(start_row=8, end_row=8,
                       start_column=col, end_column=col + 2)
        lc = ws.cell(row=8, column=col, value=lbl + " 중량")
        lc.font = Font(bold=True, size=9, color="444444")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = fill("CFD8DC")
        lc.border = thin_border()

        # 중량 수치 (row 9)
        ws.merge_cells(start_row=9, end_row=9,
                       start_column=col, end_column=col + 2)
        vc = ws.cell(row=9, column=col, value=w_formula)
        vc.font = Font(bold=True, size=14)
        vc.number_format = '0.00 "kg"'
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = fill(bg)
        vc.border = thin_border()

        # 사용률 라벨 (row 10)
        ws.merge_cells(start_row=10, end_row=10,
                       start_column=col, end_column=col + 2)
        lc2 = ws.cell(row=10, column=col, value=lbl + " 사용률")
        lc2.font = Font(bold=True, size=9, color="444444")
        lc2.alignment = Alignment(horizontal="center", vertical="center")
        lc2.fill = fill("CFD8DC")
        lc2.border = thin_border()

        # 사용률 수치 (row 11)
        ws.merge_cells(start_row=11, end_row=11,
                       start_column=col, end_column=col + 2)
        pct = ws.cell(row=11, column=col, value=pct_formula)
        pct.font = Font(bold=True, size=13)
        pct.number_format = '0.0%'
        pct.alignment = Alignment(horizontal="center", vertical="center")
        pct.fill = fill("ECEFF1")
        pct.border = thin_border()

        col_letter = get_column_letter(col)
        cell_ref_pct = f"{col_letter}11"
        ws.conditional_formatting.add(cell_ref_pct,
            CellIsRule(operator="greaterThan", formula=["1"], fill=fill(CLR["red_bad"])))
        ws.conditional_formatting.add(cell_ref_pct,
            CellIsRule(operator="between", formula=["0.9", "1"], fill=fill(CLR["orange_caution"])))
        ws.conditional_formatting.add(cell_ref_pct,
            CellIsRule(operator="lessThan", formula=["0.9"], fill=fill(CLR["green_ok"])))

    # ── 4. 균형 판정 섹션
    ws.row_dimensions[13].height = 24
    ws.merge_cells("B13:M13")
    s3 = ws["B13"]
    s3.value = "균형 판정"
    s3.font = Font(bold=True, color=CLR["white"], size=12)
    s3.fill = fill("455A64")
    s3.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[14].height = 40

    # 카약1 균형 (B:F merge)
    ws.merge_cells("B14:F14")
    k1b = ws["B14"]
    k1b.value = f"=적재_계획!B{plan_refs['k1_balance_row']}"
    k1b.font = Font(bold=True, size=22)
    k1b.alignment = Alignment(horizontal="center", vertical="center")
    k1b.border = medium_border()

    ws.conditional_formatting.add("B14:F14",
        FormulaRule(formula=['$B14="✅ 양호"'], fill=fill(CLR["green_ok"])))
    ws.conditional_formatting.add("B14:F14",
        FormulaRule(formula=['$B14="⚠️ 주의"'], fill=fill(CLR["orange_caution"])))
    ws.conditional_formatting.add("B14:F14",
        FormulaRule(formula=['$B14="❌ 불균형"'], fill=fill(CLR["red_bad"])))

    # 카약2 균형 (G:M merge)
    ws.merge_cells("G14:M14")
    k2b = ws["G14"]
    k2b.value = f"=적재_계획!D{plan_refs['k2_balance_row']}"
    k2b.font = Font(bold=True, size=22)
    k2b.alignment = Alignment(horizontal="center", vertical="center")
    k2b.border = medium_border()

    ws.conditional_formatting.add("G14:M14",
        FormulaRule(formula=['$G14="✅ 양호"'], fill=fill(CLR["green_ok"])))
    ws.conditional_formatting.add("G14:M14",
        FormulaRule(formula=['$G14="⚠️ 주의"'], fill=fill(CLR["orange_caution"])))
    ws.conditional_formatting.add("G14:M14",
        FormulaRule(formula=['$G14="❌ 불균형"'], fill=fill(CLR["red_bad"])))

    # ── 5. 총 적재 현황
    ws.row_dimensions[16].height = 24
    ws.merge_cells("B16:M16")
    s4 = ws["B16"]
    s4.value = "총 적재 현황"
    s4.font = Font(bold=True, color=CLR["white"], size=12)
    s4.fill = fill("455A64")
    s4.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[17].height = 20
    ws.row_dimensions[18].height = 26

    ds = plan_refs["data_start"]
    de = plan_refs["data_end"]

    totals = [
        ("카약1 장비수",  f'=COUNTIF(적재_계획!F{ds}:F{de},"선수")+COUNTIF(적재_계획!F{ds}:F{de},"선미")', ""),
        ("카약1 총중량",  f"=(적재_계획!B{wr}+적재_계획!C{wr})", '0.00 "kg"'),
        ("카약2 장비수",  f'=COUNTIF(적재_계획!G{ds}:G{de},"선수")+COUNTIF(적재_계획!G{ds}:G{de},"선미")', ""),
        ("카약2 총중량",  f"=(적재_계획!D{wr}+적재_계획!E{wr})", '0.00 "kg"'),
    ]

    for idx, (lbl, formula, fmt) in enumerate(totals):
        col = 2 + idx * 3
        ws.merge_cells(start_row=17, end_row=17,
                       start_column=col, end_column=col + 2)
        lc = ws.cell(row=17, column=col, value=lbl)
        lc.font = Font(bold=True, size=9, color="444444")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.fill = fill("CFD8DC")
        lc.border = thin_border()

        ws.merge_cells(start_row=18, end_row=18,
                       start_column=col, end_column=col + 2)
        vc = ws.cell(row=18, column=col, value=formula)
        vc.font = Font(bold=True, size=14)
        if fmt:
            vc.number_format = fmt
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill = fill("ECEFF1")
        vc.border = thin_border()

    # ── 6. 차트용 숨김 데이터 (rows 20~22)
    ws.row_dimensions[20].height = 0   # 숨김
    ws.cell(row=20, column=2, value="카약1선수")
    ws.cell(row=20, column=3, value="카약1선미")
    ws.cell(row=20, column=4, value="카약2선수")
    ws.cell(row=20, column=5, value="카약2선미")

    ws.cell(row=21, column=1, value="적재중량(kg)")
    ws.cell(row=21, column=2, value=f"=적재_계획!B{wr}")
    ws.cell(row=21, column=3, value=f"=적재_계획!C{wr}")
    ws.cell(row=21, column=4, value=f"=적재_계획!D{wr}")
    ws.cell(row=21, column=5, value=f"=적재_계획!E{wr}")

    ws.cell(row=22, column=1, value="허용중량(kg)")
    ws.cell(row=22, column=2, value="=카약_설정!C5")
    ws.cell(row=22, column=3, value="=카약_설정!C8")
    ws.cell(row=22, column=4, value="=카약_설정!F5")
    ws.cell(row=22, column=5, value="=카약_설정!F8")

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "선수 / 선미 중량 현황 (4그룹)"
    chart.y_axis.title = "중량 (kg)"
    chart.style = 10
    chart.width = 24
    chart.height = 14

    data_ref = Reference(ws, min_col=2, max_col=5, min_row=20, max_row=22)
    cats = Reference(ws, min_col=1, max_col=1, min_row=21, max_row=22)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "B19")


# ══════════════════════════════════════════════════════════════════════════════
# VBA 코드 출력
# ══════════════════════════════════════════════════════════════════════════════
VBA_CODE = r"""
' ================================================================================
' kayak_vba.bas  —  카약 2대 적재 계획 유틸리티 매크로
' ================================================================================
'
' ※ 적재_계획 F/G열은 이제 수식으로 자동 반영됩니다.
'    체크박스 클릭 → H/I/J/K TRUE/FALSE 변경
'    → 적재_계획 F열 =IF(장비_DB!H{n},"선수",IF(장비_DB!I{n},"선미",""))
'    → 적재_계획 G열 =IF(장비_DB!J{n},"선수",IF(장비_DB!K{n},"선미",""))
'    VBA 없이도 즉시 반영됩니다.
'
' [선택적 VBA 설치 — 전체 초기화 버튼이 필요할 때만]
'   1. Alt+F11 → VBA 에디터
'   2. 삽입 → 모듈 → Module1에 아래 코드 붙여넣기
' ================================================================================


' ── 표준 모듈 (Module1) ──────────────────────────────────────────────────────

' 전체 초기화: 체크박스(H:K)를 모두 FALSE로 → F/G 수식이 자동으로 빈칸이 됨
Sub ClearAllPlan()
    If MsgBox("모든 배치를 초기화합니다. 계속하시겠습니까?", vbYesNo) <> vbYes Then Exit Sub
    ThisWorkbook.Sheets("장비_DB").Range("H2:K51").Value = False
    MsgBox "초기화 완료 — 적재_계획이 자동 갱신됩니다.", vbInformation
End Sub


' 선수/선미 중복 방지 (같은 카약에 선수+선미 동시 체크 불가)
' → 체크박스에 직접 매크로 연결 방식으로 사용
'   개발 도구 → 삽입 → 양식 컨트롤 체크박스 우클릭 → "매크로 지정"
'   각 행/열에 맞는 Sub를 지정하거나, 아래 범용 Sub 하나를 모든 체크박스에 지정
Sub MutualExclude()
    ' 현재 선택된 체크박스의 LinkedCell을 기준으로 상호 배제
    Dim ws As Worksheet
    Set ws = ThisWorkbook.Sheets("장비_DB")
    Dim cb As Object
    Set cb = ws.Shapes(Application.Caller)
    Dim lnkCell As Range
    Set lnkCell = ws.Range(cb.ControlFormat.LinkedCell)
    If lnkCell.Value = False Then Exit Sub   ' 체크 해제 시 무시
    Dim r As Long, c As Long
    r = lnkCell.Row
    c = lnkCell.Column
    Select Case c
        Case 8:  ws.Cells(r, 9).Value = False   ' K1선수 체크 → K1선미 해제
        Case 9:  ws.Cells(r, 8).Value = False   ' K1선미 체크 → K1선수 해제
        Case 10: ws.Cells(r, 11).Value = False  ' K2선수 체크 → K2선미 해제
        Case 11: ws.Cells(r, 10).Value = False  ' K2선미 체크 → K2선수 해제
    End Select
End Sub
"""


# ══════════════════════════════════════════════════════════════════════════════
# VML 체크박스 주입 (xlsx post-processing)
# ══════════════════════════════════════════════════════════════════════════════
def inject_db_checkboxes(filename: str) -> None:
    """
    openpyxl이 저장한 xlsx에 Form Control 체크박스를 VML로 직접 주입한다.
    대상: 장비_DB 시트 H2:K51 (50행 × 4열 = 200개)
    LinkedCell: H{n}=카약1선수, I{n}=카약1선미, J{n}=카약2선수, K{n}=카약2선미
    """
    import zipfile, shutil, re

    CB_COL0S    = [6, 7, 8, 9]         # H/I/J/K 0-based 열 인덱스
    COL_LETTERS = ['H', 'I', 'J', 'K']
    CB_ROW0S    = range(0, DB_ROW_COUNT)  # DB_ROW_COUNT 기준 자동 확장

    # ── 1. 장비_DB 시트 파일 경로 찾기 ──────────────────────────────
    with zipfile.ZipFile(filename) as zf:
        wb_xml  = zf.read('xl/workbook.xml').decode('utf-8')
        wb_rels = zf.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        existing_files = set(zf.namelist())

    m = re.search(r'<sheet\b[^>]*name="장비_DB"[^>]*r:id="([^"]+)"', wb_xml)
    if not m:
        raise RuntimeError("workbook.xml에서 장비_DB 시트를 찾지 못했습니다.")
    rid = m.group(1)

    # Id 속성이 Target 앞뒤 어느 위치에 있어도 매칭
    m2 = re.search(rf'<Relationship\b([^>]*)Id="{rid}"([^>]*)>', wb_rels)
    if not m2:
        raise RuntimeError(f"workbook.xml.rels에서 {rid} 관계를 찾지 못했습니다.")
    elem_attrs = m2.group(1) + m2.group(2)
    m_target = re.search(r'Target="([^"]+)"', elem_attrs)
    if not m_target:
        raise RuntimeError(f"{rid} 관계에서 Target 속성을 찾지 못했습니다.")
    target_raw = m_target.group(1)           # e.g. "/xl/worksheets/sheet2.xml"
    # zip 내 경로는 leading '/' 없이 사용
    sheet_path     = target_raw.lstrip('/')  # "xl/worksheets/sheet2.xml"
    sheet_xml_name = target_raw.split('/')[-1]  # "sheet2.xml"
    rels_path        = f'xl/worksheets/_rels/{sheet_xml_name}.rels'
    vml_path         = 'xl/drawings/vmlDrawing_db.vml'
    vml_rid          = 'rIdVmlDb1'

    # ── 2. VML 생성 ──────────────────────────────────────────────────
    # Anchor: leftCol, leftDx, topRow, topDy, rightCol, rightDx, bottomRow, bottomDy
    # Dx/Dy: 셀 경계로부터의 오프셋 (단위: points×scale, 작은 값으로 여백)
    shapes_xml = []
    shape_id   = 1025
    z_order    = 1

    for row0 in CB_ROW0S:
        for col0, col_letter in zip(CB_COL0S, COL_LETTERS):
            row1 = row0 + 2  # 1-based (anchor는 0-based offset, 헤더행 1개 보정)
            anchor = f"{col0}, 382, {row0}, 43, {col0+1}, -382, {row0+1}, -43"
            shapes_xml.append(
                f' <v:shape id="_x0000_s{shape_id}" type="#_x0000_t201"\n'
                f'  style=\'position:absolute;margin-left:0;margin-top:0;'
                f'width:16pt;height:16pt;z-index:{z_order}\'\n'
                f'  filled="f" stroked="f" strokecolor="window [64]">\n'
                f'  <v:fill color2="window [65]" o:detectmouseclick="t"/>\n'
                f'  <v:shadow color="windowText [64]" obscured="t"/>\n'
                f'  <v:path o:connecttype="none"/>\n'
                f'  <v:textbox style=\'mso-direction-alt:auto\'>'
                f'<div style=\'text-align:left\'></div></v:textbox>\n'
                f'  <x:ClientData ObjectType="Checkbox">\n'
                f'   <x:Anchor>{anchor}</x:Anchor>\n'
                f'   <x:PrintObject/>\n'
                f'   <x:AutoFill>False</x:AutoFill>\n'
                f'   <x:FmlaLink>${col_letter}${row1}</x:FmlaLink>\n'
                f'   <x:TextHAlign>Left</x:TextHAlign>\n'
                f'   <x:FmlaGroup>0</x:FmlaGroup>\n'
                f'  </x:ClientData>\n'
                f' </v:shape>'
            )
            shape_id += 1
            z_order  += 1

    vml_content = (
        '<xml xmlns:v="urn:schemas-microsoft-com:vml"\n'
        ' xmlns:o="urn:schemas-microsoft-com:office:office"\n'
        ' xmlns:x="urn:schemas-microsoft-com:office:excel">\n'
        ' <o:shapelayout v:ext="edit">\n'
        '  <o:idmap v:ext="edit" data="1"/>\n'
        ' </o:shapelayout>\n'
        ' <v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"\n'
        '              path="m,l,21600r21600,xe">\n'
        '  <v:stroke joinstyle="miter"/>\n'
        '  <v:path shadowok="f" o:extrusionok="f" gradientshapeok="t"'
        ' o:connecttype="rect"/>\n'
        ' </v:shapetype>\n'
        + '\n'.join(shapes_xml) + '\n'
        '</xml>\n'
    )

    # ── 3. xlsx 패치 ─────────────────────────────────────────────────
    tmp = filename + '.tmp'
    with zipfile.ZipFile(filename, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                fn   = item.filename

                if fn == '[Content_Types].xml':
                    ct = data.decode('utf-8')
                    if 'vmlDrawing' not in ct:
                        ct = ct.replace(
                            '</Types>',
                            '<Default Extension="vml" ContentType="application/'
                            'vnd.openxmlformats-officedocument.vmlDrawing"/>\n'
                            '</Types>'
                        )
                    data = ct.encode('utf-8')

                elif fn == sheet_path:
                    ws_xml = data.decode('utf-8')
                    if '<legacyDrawing' not in ws_xml:
                        # r: 네임스페이스가 없으면 루트 태그에 추가
                        R_NS = 'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
                        if 'xmlns:r=' not in ws_xml:
                            ws_xml = ws_xml.replace(
                                '<worksheet ',
                                f'<worksheet {R_NS} ',
                                1
                            )
                        ws_xml = ws_xml.replace(
                            '</worksheet>',
                            f'<legacyDrawing r:id="{vml_rid}"/></worksheet>'
                        )
                    data = ws_xml.encode('utf-8')

                elif fn == rels_path:
                    rels = data.decode('utf-8')
                    if vml_rid not in rels:
                        rels = rels.replace(
                            '</Relationships>',
                            f' <Relationship Id="{vml_rid}"\n'
                            '  Type="http://schemas.openxmlformats.org/officeDocument'
                            '/2006/relationships/vmlDrawing"\n'
                            '  Target="../drawings/vmlDrawing_db.vml"/>\n'
                            '</Relationships>'
                        )
                    data = rels.encode('utf-8')

                zout.writestr(item, data)

            # 시트 rels 파일이 없으면 새로 생성
            if rels_path not in existing_files:
                new_rels = (
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org'
                    '/package/2006/relationships">\n'
                    f' <Relationship Id="{vml_rid}"\n'
                    '  Type="http://schemas.openxmlformats.org/officeDocument'
                    '/2006/relationships/vmlDrawing"\n'
                    '  Target="../drawings/vmlDrawing_db.vml"/>\n'
                    '</Relationships>\n'
                )
                zout.writestr(rels_path, new_rels.encode('utf-8'))

            # VML 파일 추가
            zout.writestr(vml_path, vml_content.encode('utf-8'))

    shutil.move(tmp, filename)
    total = len(CB_COL0S) * len(list(CB_ROW0S))
    print(f"   체크박스 {total}개 VML 주입 완료 → {vml_path}")


# ══════════════════════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet_settings(wb)
    build_sheet_db(wb)
    plan_refs = build_sheet_plan(wb)
    build_sheet_dashboard(wb, plan_refs)

    # 시트 탭 색상
    wb["카약_설정"].sheet_properties.tabColor = "1565C0"
    wb["장비_DB"].sheet_properties.tabColor   = "2E7D32"
    wb["적재_계획"].sheet_properties.tabColor  = "E65100"
    wb["대시보드"].sheet_properties.tabColor   = "6A1B9A"

    # 뷰 설정
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = True
        sheet.freeze_panes = sheet["A2"] if sheet.title != "대시보드" else None
    wb["적재_계획"].freeze_panes = "A4"

    xlsx_name = "kayak_packing_planner.xlsx"
    wb.save(xlsx_name)
    print(f"✅  {xlsx_name} 생성 완료")
    print(f"   시트: {', '.join(ws.title for ws in wb.worksheets)}")
    print(f"   장비_DB: {DB_ROW_COUNT}개 항목  /  적재_계획: 행 {PLAN_DATA_START}~{PLAN_DATA_END}")

    # VML 체크박스 주입
    inject_db_checkboxes(xlsx_name)

    # VBA 코드 파일 출력
    vba_name = "kayak_vba.bas"
    with open(vba_name, "w", encoding="utf-8") as f:
        f.write(VBA_CODE.lstrip("\n"))
    print(f"✅  {vba_name} 생성 완료")


if __name__ == "__main__":
    main()
